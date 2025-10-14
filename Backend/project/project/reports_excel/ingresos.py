
# Modelos
from apps.accountings.models import Egress, Income

# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json


from django.db.models import Q, Sum

# Tiempo
from datetime import datetime

# template
from django.views.generic import TemplateView
from django.http import HttpResponse


# Decoradores
from project.decorador import permiso_requerido
from django.utils.decorators import method_decorator


class ReporteIngresosExcelView(TemplateView):

    def get_queryset(self):
        try:
            sucursal = self.request.session['sucursal_id']
            query = self.query()
            mes = self.mes_reporte()
            anio = self.anio_reporte()

            filters = Q()

            # Filtro por texto
            if query or query != '':
                filters |= Q(fecha__icontains=query)
                filters |= Q(codigo_ingreso__icontains=query)
                filters |= Q(numero_referencia__icontains=query)

                if query.isdigit():
                    filters |= Q(monto__exact=query)

            # Filtro por mes
            if mes:
                filters &= Q(fecha__month=mes)

            # Filtro por año
            if anio:
                filters &= Q(fecha__year=anio)

            return Income.objects.filter(filters, sucursal=sucursal)
        except Exception as e:
            print(f"Error al filtrar el queryset: {e}")
            return Income.objects.none()

    

    def query(self):
        return self.request.GET.get('q')
    
    def mes_reporte(self):
        return self.request.GET.get('mes')
    
    def anio_reporte(self):
        return self.request.GET.get('anio')


    def get(self, request, *args, **kwargs):
        # Crear el archivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Ingresos"

        # Encabezados
        encabezados = [
            "#", "CODIGO DE INGRESOS", "FECHA",  "NUMERO DE REFERENCIA", "MONTO", "DESCRIPCION",
            "OBSERVACIONES"
        ]
        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Obtener datos
        informacion_consultada = self.get_queryset()

        contador = 0

        for idx, ingreso in enumerate(informacion_consultada, start=2):
            contador += 1
            
            fila = [
                contador,
                ingreso.codigo_ingreso,
                ingreso.fecha,
                ingreso.numero_referencia,
                ingreso.monto,
                ingreso.descripcion,
                ingreso.observaciones
            ]

            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)

        # Crear respuesta de descarga
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_ingresos_{timestamp}.xlsx"'
        workbook.save(response)
        return response

