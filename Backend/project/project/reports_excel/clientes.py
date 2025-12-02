
from openpyxl import Workbook
from django.http import HttpResponse
import json

from django.db.models import Q, Sum
from django.views.generic import TemplateView
from datetime import datetime

# MODELOS
from apps.customers.models import Customer, CreditCounselor
from apps.financings.models import Credit, PaymentPlan, Disbursement, Guarantees
from apps.addresses.models import Address
from apps.FinancialInformation.models import WorkingInformation, Reference

def informacion_credito(creditos):
    return creditos.first() if creditos.exists() else None


def obtener_fiadores(reporte):

    if not reporte.es_fiador():
        return 'No Tiene'
    
    resultado = ''

    for fiador in reporte.es_fiador():
        resultado +=f'{fiador.garantia_id.credit_id} '

    return resultado
 
def report_clientes(request):
    # Crear el archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de cartera de Creditos"

    # Agregar encabezados
    sheet['A1'] = "#"
    sheet['B1'] = "NOMBRE"
    sheet['C1'] = "TIPO DE IDENTIFICACION"
    sheet['D1'] = "NUMERO DE IDENTIFICACION"
    sheet['E1'] = "NUMERO DE TELEFONO"
    sheet['F1'] = "EDAD"
    sheet['G1'] = "GENERO"
    sheet['H1'] = "CODIGO DE CLIENTE"
    sheet['I1'] = "PROFESION U OFICIO"
    sheet['J1'] = "ASESOR DEL CREDITO"
    sheet['K1'] = "TIENE CREDITO"
    sheet['L1'] = "CANTIDAD DE CREDITOS"
    sheet['M1'] = "CODIGO DE CREDITO"
    sheet['N1'] = "EL CREDITO ESTA VIGENTE"
    sheet['O1'] = "PROPOSITO"
    sheet['P1'] = "MONTO"
    sheet['Q1'] = "PLAZO"
    sheet['R1'] = "TASA DE INTERES"
    sheet['S1'] = "SALDO CAPITAL PENDIENTE"
    sheet['T1'] = "SALDO ACTUAL"
    sheet['U1'] = "FECHA DE INICIO DEL CREDITO"
    sheet['V1'] = "FECHA DE VENCIMIENTO DEL CREDITO"
    sheet['W1'] = "FECHA DE INICIO DE LA CUOTA"
    sheet['X1'] = "FECHA DE VENCIMIENTO DE LA CUOTA"
    sheet['Y1'] = "FECHA LIMITE DE LA CUOTA"
    sheet['Z1'] = "TIPO DE CREDITO"
    sheet['AA1'] = "FORMA DE PAGO"
    sheet['AB1'] = "ESTADO POR FECHAS"
    sheet['AC1'] = "ESTADO POR APORTACION"
    sheet['AD1'] = 'FORMA DE DESEMBOLSO'
    sheet['AE1'] = 'TIPO DE GARANTIA'

    sheet['AF1'] = 'DIRECCION DEL CLIENTE'
    sheet['AG1'] = 'MUNICIPIO DEL CLIENTE' 
    sheet['AH1'] = 'DEPARTAMENTO DEL CLIENTE' 
    sheet['AI1'] = 'DIRECCION DE TRABAJO'
    sheet['AJ1'] = 'MUNICIPIO DE LABORAL' 
    sheet['AK1'] = 'DEPARTAMENTO DE LABORAL' 

    sheet['AL1'] = 'FUENTE DE INGRESO'
    sheet['AM1'] =  'ESTADO LABORAL'
    sheet['AN1'] = 'EMPRESA DE LABORAL'
    sheet['AO1'] =  'PUESTO'

    sheet['AP1'] = 'REFERENCIA 1'
    sheet['AQ1'] = 'REFERENCIA 2'
    sheet['AR1'] = 'REFERENCIA 3'
    sheet['AS1'] = 'REFERENCIA 4'

    


    # Obtener datos de la base de datos, ordenando por status (True primero)
    from django.db.models import Count

    # Obtener datos de la base de datos, ordenando para mostrar primero los que tienen créditos
    clientes = Customer.objects.annotate(num_creditos=Count('credit')).order_by('-num_creditos')

    contador = 0

    # Agregar datos al archivo
    for idx, cliente in enumerate(clientes, start=2):  # Comenzar en la fila 2
        contador += 1
        creditos = Credit.objects.filter(customer_id__id=cliente.id).order_by('-id')
        
        primer_credito = creditos.first()

        cuota = None
        desembolso = None
        garantia = None

        


        mensaje = ''
        moroso = ''
        tipo_garantia = ''

        informacion_laboral = WorkingInformation.objects.filter(customer_id=cliente.id).first()
        direcion_personal = Address.objects.filter(customer_id=cliente.id ).exclude(type_address='Dirección de Trabajo').first()
        filters = Q()
        filters |= Q(type_address='Dirección de Trabajo')
        filters |= Q(type_address='Direccin de Trabajo')

        direccion_laboral = Address.objects.filter(Q(customer_id=cliente.id), filters).first()
        referencias = Reference.objects.filter(customer_id=cliente.id)


        

        if primer_credito:
            mensaje = 'CREDITO CANCELADO' if primer_credito.is_paid_off else 'CREDITO VIGENTE'
            moroso = 'NO' if primer_credito.estados_fechas else 'SI'
            cuota = PaymentPlan.objects.filter(credit_id= primer_credito.id).order_by('-id').first()
            desembolso = Disbursement.objects.filter(credit_id=primer_credito.id).exclude(forma_desembolso='DESEMBOLSAR').first()
            garantia =  Guarantees.objects.filter(credit_id = primer_credito.id).first()
            

        sheet[f'A{idx}'] = contador
        sheet[f'B{idx}'] = str(cliente)
        sheet[f'C{idx}'] = str(cliente.type_identification)
        sheet[f'D{idx}'] = str(cliente.identification_number)
        sheet[f'E{idx}'] = cliente.telephone
        sheet[f'F{idx}'] = str(cliente.get_edad())
        sheet[f'G{idx}'] = cliente.gender
        sheet[f'H{idx}'] = cliente.customer_code
        sheet[f'I{idx}'] = cliente.profession_trade
        sheet[f'J{idx}'] = cliente.asesor
        sheet[f'K{idx}'] = 'SI' if creditos else 'NO CUENTA CON CREDITOS REGISTRADOS'
        sheet[f'L{idx}'] =  str(creditos.count()) if creditos else '0'
        sheet[f'M{idx}'] = str(primer_credito.codigo_credito) if primer_credito else ''
        sheet[f'N{idx}'] = mensaje if primer_credito else ''
        sheet[f'O{idx}'] = primer_credito.proposito if primer_credito else ''
        sheet[f'P{idx}'] = primer_credito.monto if primer_credito else ''
        sheet[f'Q{idx}'] = primer_credito.plazo if primer_credito else ''
        sheet[f'R{idx}'] = primer_credito.tasa_mensual() if primer_credito else ''
        sheet[f'S{idx}'] = primer_credito.formato_saldo_pendiente() if primer_credito else ''
        sheet[f'T{idx}'] = primer_credito.formato_saldo_actual() if primer_credito else ''
        sheet[f'U{idx}'] = str(primer_credito.fecha_inicio) if primer_credito else ''
        sheet[f'V{idx}'] = str(primer_credito.fecha_vencimiento) if primer_credito else ''
        sheet[f'W{idx}'] = str(cuota.start_date.date()) if cuota  else ''
        sheet[f'X{idx}'] = str(cuota.due_date.date()) if cuota  else ''
        sheet[f'Y{idx}'] = str(cuota.mostrar_fecha_limite().date()) if cuota else ''
        sheet[f'Z{idx}'] = primer_credito.tipo_credito if primer_credito else ''
        sheet[f'AA{idx}'] = primer_credito.forma_de_pago if primer_credito else ''
        sheet[f'AB{idx}'] = primer_credito.formato_estado_fecha() if primer_credito else ''
        sheet[f'AC{idx}'] = primer_credito.formato_estado_aportacion() if primer_credito else ''
        sheet[f'AD{idx}'] = desembolso.forma_desembolso if desembolso else ''
        sheet[f'AE{idx}'] = garantia.tipos_garantia() if garantia else ''
        
        
        sheet[f'AF{idx}'] = direcion_personal.street if direcion_personal else ''
        sheet[f'AG{idx}'] = direcion_personal.state if direcion_personal else ''
        sheet[f'AH{idx}'] = direcion_personal.city if direcion_personal else ''
        sheet[f'AI{idx}'] = direccion_laboral.street  if direccion_laboral else ''
        sheet[f'AJ{idx}'] = direccion_laboral.state  if direccion_laboral else ''
        sheet[f'AK{idx}'] = direccion_laboral.city if direccion_laboral else ''
        
        sheet[f'AL{idx}'] = informacion_laboral.get_fuente_ingreso() if informacion_laboral else ''
        sheet[f'AM{idx}'] = informacion_laboral.get_estado_laboral() if informacion_laboral else ''
        sheet[f'AN{idx}'] = informacion_laboral.get_empresa_laburo() if informacion_laboral else ''
        sheet[f'AO{idx}'] = informacion_laboral.get_puesto() if informacion_laboral else ''
        
        if referencias:
            columnas = ['AP', 'AQ', 'AR', 'AS']

            for i, referencia in enumerate(referencias[:4]):  # Limita a máximo 4 referencias
                sheet[f'{columnas[i]}{idx}'] = referencia.full_name
    

    

       




    # Crear respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="reporte_cartera_creditos_{datetime.now()}.xlsx"'

    # Guardar el archivo en la respuesta
    workbook.save(response)
    return response


class ReporteClientesExcelView(TemplateView):
    def get_queryset(self):
        try:
            # Asignar la consulta a una variable local
            query = self.query()
            gender = self.genero()
            incompleto = self.falta_informacion()
            status = self.status()
            sucursal = self.sucursal()
            request = self.request

            filters = Q()

            asesor_autenticado = CreditCounselor.objects.filter(usuario=request.user).first()

            if query:                
                filters |= Q(first_name__icontains=query) 
                filters |= Q(customer_code__icontains=query)
                filters |= Q(last_name__icontains=query)
                filters |= Q(type_identification__icontains=query)
            
            if gender:
                filters |= Q(gender__icontains=gender)

            if incompleto:
                filters |= Q(completado=False)
            
            if status:
                match status:
                    case 'Falta de Informacion':
                        filters &= Q(completado=False)
                    case 'Clientes Dados de Baja':
                        filters &= Q(status__icontains='Dar de Baja')
                    case _:
                        filters &= Q(status__icontains=status)

            
            if sucursal:
                filters &= Q(sucursal__nombre__icontains=sucursal)
            
            if asesor_autenticado and request.user.rol.role_name == 'Asesor de Crédito':
                filters &= Q(new_asesor_credito=asesor_autenticado) 


 
            # Filtrar los objetos Customer usando los filtros definidos
            return Customer.objects.filter(filters).order_by('-id')
        except Exception as e:
            # Manejar cualquier excepción que ocurra y devolver un queryset vacío
            print(f"Error al filtrar el queryset: {e}")
            
            return Customer.objects.none()
    


    def query(self):
        return self.request.GET.get('q')
    
    def genero(self):
        return self.request.GET.get('gender')
    
    def falta_informacion(self):
        return self.request.GET.get('incompleto')
    
    def status(self):
        return self.request.GET.get('status')
    
    def sucursal(self):
        return self.request.GET.get('sucursal')
    
    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de cartera de Clientes"
        encabezados = [
            "#", "NOMBRE", "TIPO DE IDENTIFICACION", "NUMERO DE IDENTIFICACION", "NUMERO DE TELEFONO", "EDAD", "GENERO",
            "CODIGO DE CLIENTE", "PROFESION U OFICIO", "ASESOR DEL CREDITO", 'FIADOR DE', "TIENE CREDITO", "CANTIDAD DE CREDITOS",
            "CODIGO DE CREDITO", "EL CREDITO ESTA VIGENTE", "PROPOSITO", "MONTO", "PLAZO", "TASA DE INTERES",
            "SALDO CAPITAL PENDIENTE", "SALDO ACTUAL", "FECHA DE INICIO DEL CREDITO", "FECHA DE VENCIMIENTO DEL CREDITO",
            "FECHA DE INICIO DE LA CUOTA", "FECHA DE VENCIMIENTO DE LA CUOTA", "FECHA LIMITE DE LA CUOTA", "TIPO DE CREDITO",
            "FORMA DE PAGO", "ESTADO POR FECHAS", "ESTADO POR APORTACION", "FORMA DE DESEMBOLSO", "TIPO DE GARANTIA",
            "DIRECCION DEL CLIENTE", "MUNICIPIO DEL CLIENTE", "DEPARTAMENTO DEL CLIENTE",
            "DIRECCION DE TRABAJO", "MUNICIPIO DE LABORAL", "DEPARTAMENTO DE LABORAL",
            "FUENTE DE INGRESO", "ESTADO LABORAL", "EMPRESA DE LABORAL", "PUESTO",
            "REFERENCIA 1", "REFERENCIA 2", "REFERENCIA 3", "REFERENCIA 4", 
        ]


        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Obtener datos
        informacion_consultada = self.get_queryset()

        contador = 0
        for idx, cliente in enumerate(informacion_consultada, start=2):
            contador += 1
            creditos = Credit.objects.filter(customer_id__id=cliente.id).order_by('-id')
        
            primer_credito = creditos.first()

            cuota = None
            desembolso = None
            garantia = None

            mensaje = ''
            moroso = ''
            tipo_garantia = ''

            informacion_laboral = WorkingInformation.objects.filter(customer_id=cliente.id).first()
            direcion_personal = Address.objects.filter(customer_id=cliente.id ).exclude(type_address='Dirección de Trabajo').first()
            filters = Q()
            filters |= Q(type_address='Dirección de Trabajo')
            filters |= Q(type_address='Direccin de Trabajo')

            direccion_laboral = Address.objects.filter(Q(customer_id=cliente.id), filters).first()
            referencias = Reference.objects.filter(customer_id=cliente.id)


        

            if primer_credito:
                mensaje = 'CREDITO CANCELADO' if primer_credito.is_paid_off else 'CREDITO VIGENTE'
                moroso = 'NO' if primer_credito.estados_fechas else 'SI'
                cuota = PaymentPlan.objects.filter(credit_id= primer_credito.id).order_by('-id').first()
                desembolso = Disbursement.objects.filter(credit_id=primer_credito.id).exclude(forma_desembolso='DESEMBOLSAR').first()
                garantia =  Guarantees.objects.filter(credit_id = primer_credito.id).first()
                

            
            # Creamos una lista con todos los valores de la fila
            fila = [
                contador,
                str(cliente),
                str(cliente.type_identification),
                str(cliente.identification_number),
                cliente.telephone,
                str(cliente.get_edad()),
                cliente.gender,
                cliente.customer_code,
                cliente.profession_trade,
                cliente.asesor,
                str(obtener_fiadores(cliente)), 
                'SI' if creditos else 'NO CUENTA CON CREDITOS REGISTRADOS',
                str(creditos.count()) if creditos else '0',
                str(primer_credito.codigo_credito) if primer_credito else '',
                mensaje if primer_credito else '',
                primer_credito.proposito if primer_credito else '',
                primer_credito.monto if primer_credito else '',
                primer_credito.plazo if primer_credito else '',
                primer_credito.tasa_mensual() if primer_credito else '',
                primer_credito.formato_saldo_pendiente() if primer_credito else '',
                primer_credito.formato_saldo_actual() if primer_credito else '',
                str(primer_credito.fecha_inicio) if primer_credito else '',
                str(primer_credito.fecha_vencimiento) if primer_credito else '',
                str(cuota.start_date.date()) if cuota else '',
                str(cuota.due_date.date()) if cuota else '',
                str(cuota.mostrar_fecha_limite().date()) if cuota else '',
                primer_credito.tipo_credito if primer_credito else '',
                primer_credito.forma_de_pago if primer_credito else '',
                primer_credito.formato_estado_fecha() if primer_credito else '',
                primer_credito.formato_estado_aportacion() if primer_credito else '',
                desembolso.forma_desembolso if desembolso else '',
                garantia.tipos_garantia() if garantia else '',
                direcion_personal.street if direcion_personal else '',
                direcion_personal.state if direcion_personal else '',
                direcion_personal.city if direcion_personal else '',
                direccion_laboral.street if direccion_laboral else '',
                direccion_laboral.state if direccion_laboral else '',
                direccion_laboral.city if direccion_laboral else '',
                informacion_laboral.get_fuente_ingreso() if informacion_laboral else '',
                informacion_laboral.get_estado_laboral() if informacion_laboral else '',
                informacion_laboral.get_empresa_laburo() if informacion_laboral else '',
                informacion_laboral.get_puesto() if informacion_laboral else '',
                               
            ]

            # Agregar las referencias (hasta 4)
            for i in range(4):
                if referencias and i < len(referencias):
                    fila.append(referencias[i].full_name)
                else:
                    fila.append('')

            # Escribir automáticamente cada valor en su columna
            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)


        # Crear respuesta de descarga
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_clientes_{timestamp}.xlsx"'
        workbook.save(response)
        return response
