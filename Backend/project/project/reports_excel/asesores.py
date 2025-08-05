# Modelos
from apps.customers.models import CreditCounselor
from apps.financings.models import Credit, PaymentPlan

# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile



from django.db.models import Q, Sum

# Tiempo
from datetime import datetime, timedelta

# template
from django.views.generic import TemplateView
from django.http import HttpResponse


# Decoradores
from project.decorador import permiso_requerido
from django.utils.decorators import method_decorator

def totales(dia, acreedor):
    dia_mas_uno = dia + timedelta(days=1)
    
    mora = 0
    interes = 0
    saldo_cap = 0
    saldo_total = 0

    # Obtener las primeras cuotas activas por crédito del asesor
    creditos = PaymentPlan.objects.filter(
        credit_id__customer_id__new_asesor_credito__id=acreedor.id,
        start_date__lte=dia,
        fecha_limite__gte=dia_mas_uno,
        credit_id__is_paid_off=False
    ).order_by('credit_id', 'start_date')  # Ordena por crédito y fecha inicial

    vistos = set()  # Para evitar repetir créditos

    for cuota in creditos:
        if cuota.credit_id_id in vistos:
            continue  # Ya contamos una cuota para este crédito

        vistos.add(cuota.credit_id_id)

        mora += cuota.mora
        interes += cuota.interest
        saldo_cap += cuota.saldo_pendiente
        saldo_total += cuota.saldo_pendiente + cuota.mora + cuota.interest

    return mora, interes, saldo_cap, saldo_total



def crear_excel_creditos_por_asesor(dia,acreedor):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Créditos del Asesor"

    encabezados = [
        "#", "NOMBRE DEL CLIENTE", "CÓDIGO DEL CRÉDITO", "FECHA LIMITE","MORA", "INTERÉS", "SALDO CAPITAL", "SALDO TOTAL"
    ]
    for col_idx, header in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

 
    dia_mas_uno = dia + timedelta(days=1)

    creditos = PaymentPlan.objects.filter(
        credit_id__customer_id__new_asesor_credito__id=acreedor.id,
        start_date__lte=dia,
        fecha_limite__gte=dia_mas_uno,
        credit_id__is_paid_off=False
    ).order_by('credit_id', 'start_date')  # Ordenar para que la primera cuota sea la más temprana

    vistos = set()  # IDs de créditos ya procesados

    contador = 0  # Para la numeración real del archivo

    for idx, credito in enumerate(creditos, start=2):
        credit_id = credito.credit_id_id

        if credit_id in vistos:
            continue  # Ya procesamos una cuota para este crédito

        vistos.add(credit_id)
        contador += 1

        nombre = f'{credito.credit_id.customer_id.first_name} {credito.credit_id.customer_id.last_name}'
        codigo = credito.credit_id.codigo_credito
        mora = credito.mora
        interes = credito.interest
        saldo_cap = credito.saldo_pendiente
        saldo_total = mora + interes + saldo_cap
        fecha_limite = credito.fecha_limite.date()

        fila = [contador, nombre, codigo, fecha_limite, mora, interes, saldo_cap, saldo_total]

        for col_idx, value in enumerate(fila, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)


    return workbook





class ReporteAsesoresExcelView(TemplateView):

    def get_queryset(self):
        try:
            # Asignar la consulta a una variable local
            query = self.query()

            filters = Q()
            fecha_consulta = datetime.now().date()  # Por defecto, hoy

            if query:
                try:
                    # Si es una fecha válida, sobrescribe fecha_consulta
                    fecha_consulta = datetime.strptime(query, '%Y-%m-%d').date()
                except ValueError:
                    # Si no es fecha válida, intenta otros filtros textuales
                    filters |= Q(nombre__icontains=query)
                    filters |= Q(apellido__icontains=query)
                    filters |= Q(codigo_asesor__icontains=query)
                    filters |= Q(identification_number__icontains=query)
                    filters |= Q(tipo_pago__icontains=query)

            # Filtrar asesores
            asesor_consulta = CreditCounselor.objects.filter(filters).order_by('-id')
            return fecha_consulta, asesor_consulta

        except Exception as e:
            print(f"Error al filtrar el queryset: {e}")
            fecha_fallback = datetime.now().date()
            asesor_consulta = CreditCounselor.objects.all().order_by('-id')
            return fecha_fallback, asesor_consulta

        
    def query(self):
        return self.request.GET.get('q')
    
    def get(self, request, *args, **kwargs):
        zip_buffer = io.BytesIO()
        zip_file = zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED)

        # 1. Crear el Excel general de acreedores
        workbook_main = Workbook()
        sheet = workbook_main.active
        sheet.title = "Reporte sobre Saldos de Asesores"

        encabezados = [
            "#", "NOMBRE", "MORA TOTAL", "INTERES TOTAL", "SALDO TOTAL DE CAPITAL PENDIENTE", "SALDO TOTAL ACTUAL"
        ]
        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        fecha,acreedores = self.get_queryset()
        contador = 0

        for idx, acreedor in enumerate(acreedores, start=2):
            contador += 1
            nombre = f'{acreedor.nombre} {acreedor.apellido}'
            mora, interes, saldo_cap, saldo_total = totales(fecha,acreedor)

            fila = [contador, nombre, mora, interes, saldo_cap, saldo_total]
            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)

            # 2. Crear Excel individual por asesor y agregarlo al zip
            asesor_buffer = io.BytesIO()
            wb_asesor = crear_excel_creditos_por_asesor(fecha,acreedor)
            wb_asesor.save(asesor_buffer)
            asesor_buffer.seek(0)
            zip_file.writestr(f"reporte_creditos_{acreedor.nombre}_{acreedor.apellido}.xlsx", asesor_buffer.read())

        # 3. Guardar el Excel general en el zip
        main_excel_buffer = io.BytesIO()
        workbook_main.save(main_excel_buffer)
        main_excel_buffer.seek(0)
        zip_file.writestr("reporte_general_acreedores.xlsx", main_excel_buffer.read())

        zip_file.close()

        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_de_saldos_de_asesores_{timestamp}_fecha_{fecha}.zip"'
        return response
