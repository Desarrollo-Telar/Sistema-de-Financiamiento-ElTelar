
from openpyxl import Workbook
from django.http import HttpResponse
import json

from django.db.models import Q, Sum
from datetime import datetime, timedelta


# Modelos
from apps.financings.models import Credit, PaymentPlan, Disbursement
from apps.customers.models import CreditCounselor

from django.views.generic import TemplateView
from django.shortcuts import render, get_object_or_404, redirect

def cuota(credito):
    dia = datetime.now().date()
    dia_mas_uno = dia + timedelta(days=1)
    
    siguiente_pago = PaymentPlan.objects.filter(
        credit_id=credito,
        start_date__lte=dia,
        fecha_limite__gte=dia_mas_uno
    ).first()

    if siguiente_pago is None:
        siguiente_pago = PaymentPlan.objects.filter(
        credit_id=credito).order_by('-id').first()
    
    return siguiente_pago

def report_creditos(request, filtro_seleccionado):
    hoy = datetime.now()
    inicio = hoy - timedelta(days=5)
    sucursal = request.session['sucursal_id']

    # Diccionario de filtros
    filtros = {
        'Todos': lambda: Credit.objects.filter(sucursal=sucursal),
        'Recientes': lambda: Credit.objects.filter(creation_date__range=[inicio, hoy], sucursal=sucursal),
        'Creditos Cancelados': lambda: Credit.objects.filter(is_paid_off=True, sucursal=sucursal),
        'Creditos en Atraso': lambda: Credit.objects.filter(estados_fechas=False, sucursal=sucursal),
        'Creditos con falta de Aportacion': lambda: Credit.objects.filter(estado_aportacion=False, sucursal=sucursal),
        'Creditos con excedente': lambda: Credit.objects.filter(Q(saldo_actual__lt=0) | Q(excedente__gt=0), sucursal=sucursal),
    }

    # Obtener los créditos según el filtro seleccionado
    reportes = filtros.get(filtro_seleccionado, lambda: Credit.objects.none())().order_by('-id')
    
    # Crear el archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Reporte de CREDITOS"

    # Agregar encabezado fecha_entrar_en_mora
    sheet['A1'] = f'REPORTE SOBRE CREDITOS'
    sheet.append([
        "#", "FECHA DE REGISTRO", "CODIGO DEL CREDITO", 
        "CLIENTE", "MONTO OTORGADO", "PROPOSITO", "PLAZO EN MESES", "TASA DE INTERES",
        "FORMA DE PAGO", "TIPO DE CREDITO", "DESEMBOLSO","FECHA DE INICIO DEL CREDITO", 
        "FECHA DE VENCIMIENTO DEL CREDITO", "FECHA LIMITE DE PAGO", "FECHA EN QUE ENTRO A MORA", 
        "SALDO ACTUAL", "SALDO CAPITAL PENDIENTE","SALDO EXCEDENTE" ,"STATUS POR FECHAS","STATUS POR APORTACION","STATUS JUDICIAL","STATUS CANCELADO", "NUMERO DE REFERENCIA", "ASESOR DE CREDITO"
    ])

    # Agregar los datos
    for idx, reporte in enumerate(reportes, start=1):
        cuota_limite = cuota(reporte.id)
        desembolso = Disbursement.objects.filter(credit_id__id=reporte.id).first()
        fecha_limite_pago = cuota_limite.mostrar_fecha_limite().date() if cuota_limite else "---"
        desembolso_forma = desembolso.forma_desembolso if desembolso else "---"
        numero_referencia = cuota_limite.numero_referencia if cuota_limite else '---'
        

        # Mensajes de estado
        if reporte.estado_aportacion:
            mensaje = 'VIGENTE'
        elif reporte.estado_aportacion is None:
            mensaje = 'SIN APORTACIONES'
        else:
            mensaje = 'EN ATRASO'

        aportacion = mensaje
        s_fecha = 'VIGENTE' if reporte.estados_fechas else 'EN ATRASO'
        stat = f'''Status de Aportación: {aportacion}, Status por Fecha: {s_fecha}'''

        s_judicial = 'Si' if reporte.estado_judicial else 'No' 
        s_cancelado = 'Si' if reporte.is_paid_off else 'No'
        sheet.append([
            idx,
            reporte.creation_date.date(),
            reporte.codigo_credito,
            reporte.customer_id.get_full_name() if reporte.customer_id else "Sin cliente",
            reporte.formato_monto(),
            reporte.proposito,
            reporte.plazo,
            reporte.tasa_mensual(),
            str(reporte.forma_de_pago),
            str(reporte.tipo_credito),
            str(desembolso_forma),
            reporte.fecha_inicio,
            reporte.fecha_vencimiento,
            fecha_limite_pago,
            str('---'),
            reporte.formato_saldo_actual(),
            reporte.formato_saldo_pendiente(),
            reporte.formato_saldo_excedente(),
            s_fecha,
            aportacion,
            s_judicial,
            s_cancelado,
            numero_referencia,
            str(reporte.customer_id.asesor)


        ])

    # Crear la respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="reportes_sobre_creditos_{filtro_seleccionado}.xlsx"'

    # Guardar el archivo en la respuesta
    workbook.save(response)
    return response

def obtener_fiadores(reporte):

    if not reporte.mi_fiador_es():
        return 'No Tiene'
    
    resultado = ''

    for fiador in reporte.mi_fiador_es():
        resultado +=f'{fiador.especificaciones['Nombre']} '

    return resultado


class ReporteCreditos(TemplateView):
    def get_queryset(self):
        try:
            request = self.request
            

            # Asignar la consulta a una variable local
            sucursal = self.request.session['sucursal_id']

            query = self.query()
            status = self.status()
            mes = self.query_mes()
            anio = self.query_anio()

            asesor_autenticado = CreditCounselor.objects.filter(usuario=request.user).first()

            # Crear una lista para almacenar los filtros
            filters = Q()

            
            
            # Añadir filtros si la consulta no está vacía
            if query:
                filters |= Q(fecha_inicio__icontains=query)
                filters |= Q(fecha_vencimiento__icontains=query)
                filters |= Q(tipo_credito__icontains=query)
                filters |= Q(proposito__icontains=query)
                filters |= Q(tipo_credito__icontains=query)
                filters |= Q(forma_de_pago__icontains=query)
                filters |= Q(codigo_credito__icontains=query)
                filters |= Q(customer_id__customer_code__icontains=query)
                filters |= Q(customer_id__first_name__icontains=query)
                filters |= Q(customer_id__last_name__icontains=query)


                # Si la consulta es numérica, usar filtro exacto para campos numéricos
                if query.isdigit():
                    filters |= Q(monto__exact=query)
                    filters |= Q(plazo__exact=query)
                    filters |= Q(tasa_interes__exact=query)
            
            if status:
                match status:
                    case 'Recientes':
                        hoy = datetime.now()
                        inicio = hoy - timedelta(days=5)
                        filters &= Q(creation_date__range=[inicio,hoy])

                    case 'Creditos Cancelados':
                        filters &= Q(is_paid_off=True)

                    case 'Creditos en Atraso':
                        filters &= Q(estados_fechas=False)
                    
                    case 'Creditos Falta de Aportacion':
                        filters &= Q(estado_aportacion=False)

                    case 'Creditos en Estado Juridico':
                        filters &= Q(estado_judicial=True)

                    case 'Creditos con Excedente':
                        filters &= (Q(saldo_actual__lt=0) | Q(excedente__gt=0))

            
            if asesor_autenticado is not None and request.user.rol.role_name == 'Asesor de Crédito':
                filters &= Q(asesor_de_credito=asesor_autenticado)

            
            
            if anio:
                filters &= Q(creation_date__year=anio)
            
            if mes:
                filters &= Q(creation_date__month=mes)
            
            if sucursal:
                filters &= Q(sucursal=sucursal)

            

            # Filtrar los objetos Banco usando los filtros definidos
            return Credit.objects.filter(filters)
        
        except Exception as e:
            # Manejar cualquier excepción que ocurra y devolver un queryset vacío
            print(f"Error al filtrar el queryset: {e}")
            return Credit.objects.none()

    def query(self):
        return self.request.GET.get('q')
    
    def status(self):
        return self.request.GET.get('status')
    
    def query_mes(self):
        return self.request.GET.get('mes')
    
    def query_anio(self):
        return self.request.GET.get('anio')
    
    def get(self, request, *args, **kwargs):
        mes = request.GET.get('mes')
        anio = request.GET.get('anio')
        query = request.GET.get('q')
        status = request.GET.get('status')

        if not query and not status:
            if mes and anio:
                return redirect('report_desmbolso', mes,anio)
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Cartera de Creditos"
        encabezados = [
            "#", "FECHA DE REGISTRO", "CODIGO DEL CREDITO", 
            "CLIENTE", "MONTO OTORGADO", "PROPOSITO", "PLAZO EN MESES", "TASA DE INTERES",
            "FORMA DE PAGO", "TIPO DE CREDITO", "DESEMBOLSO","FECHA DE INICIO DEL CREDITO", 
            "FECHA DE VENCIMIENTO DEL CREDITO", "FECHA LIMITE DE PAGO", "FECHA DE CANCELACION DEL CREDITO", "FECHA EN ENTRAR A MORA", "DIAS DE MORA",
            "SALDO ACTUAL", "SALDO CAPITAL PENDIENTE","SALDO EXCEDENTE" ,"STATUS POR FECHAS",'STATUS POR APORTACION','STATUS JUDICIAL','STATUS DEL CREDITO', "NUMERO DE REFERENCIA", "ASESOR DE CREDITO", "FIADORES DEL CREDITO",
            "RESPONSABLE DE LA GESTION", "ESTADO DE LA COBRANZA", "RESULTADO DE COBRANZA","FECHA DE SEGUIMIENTO O DE PROMESA DE PAGO"
        ]

        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Obtener datos
        informacion_consultada = self.get_queryset()
        contador = 0

        for idx, reporte in enumerate(informacion_consultada, start=2):
            contador += 1
            cuota_limite = cuota(reporte.id)
            desembolso = Disbursement.objects.filter(credit_id__id=reporte.id).first()
            fecha_limite_pago = cuota_limite.mostrar_fecha_limite().date() if cuota_limite else "---"
            desembolso_forma = desembolso.forma_desembolso if desembolso else "---"
            numero_referencia = cuota_limite.numero_referencia if cuota_limite else '---'
            

            # Mensajes de estado
            if reporte.estado_aportacion:
                mensaje = 'VIGENTE'
            elif reporte.estado_aportacion is None:
                mensaje = 'SIN APORTACIONES'
            else:
                mensaje = 'EN ATRASO'

            aportacion = mensaje
            s_fecha = 'VIGENTE' if reporte.estados_fechas else 'EN ATRASO'
            s_judicial = 'EN PROCESO JUDICIAL' if reporte.estado_judicial else 'NO'
            s_credito = 'CANCELADO' if reporte.is_paid_off else 'VIGENTE'
            responsable = reporte.tiene_gestion_cobranza().asesor_credito if reporte.tiene_gestion_cobranza() else 'NO TIENE'
            estado = reporte.tiene_gestion_cobranza().estado_cobranza if reporte.tiene_gestion_cobranza() else 'NO TIENE'
            resultado = reporte.tiene_gestion_cobranza().resultado if reporte.tiene_gestion_cobranza() else 'NO TIENE'
            fecha = reporte.tiene_gestion_cobranza().get_fecha_seg_o_promesa() if reporte.tiene_gestion_cobranza() else 'NO TIENE'

            fila = [
                contador,
                reporte.creation_date.date(),
                reporte.codigo_credito,
                reporte.customer_id.get_full_name() if reporte.customer_id else "Sin cliente",
                reporte.formato_monto(),
                reporte.proposito,
                reporte.plazo,
                reporte.tasa_mensual(),
                str(reporte.forma_de_pago),
                str(reporte.tipo_credito),
                str(desembolso_forma),
                reporte.fecha_inicio,
                reporte.fecha_vencimiento,
                fecha_limite_pago,
                reporte.fecha_cancelacion,
                reporte.fecha_entrar_en_mora,
                reporte.dias_de_mora(),
                reporte.formato_saldo_actual(),
                reporte.formato_saldo_pendiente(),
                reporte.formato_saldo_excedente(),
                s_fecha,aportacion,s_judicial,s_credito,
                numero_referencia,
                str(reporte.asesor_de_credito),
                str(obtener_fiadores(reporte)),
                str(responsable),
                str(estado),
                str(resultado),
                str(fecha)
            ]

            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)

        # Crear respuesta de descarga
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_creditos_{timestamp}.xlsx"'
        workbook.save(response)
        return response
    