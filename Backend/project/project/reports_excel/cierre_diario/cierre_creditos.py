# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile

# Tiempo
from datetime import datetime, timedelta, date

# Modelo
from apps.customers.models import Customer, CreditCounselor
from apps.financings.formato import formatear_numero

# Filtros
from .filtros.creditos import *
from .funcion import safe_list_get

def get_cliente(id):
    cliente =  Customer.objects.filter(id=id).first()

    return cliente if cliente else None

def get_asesor_credito(id):
    asesor = CreditCounselor.objects.filter(id=id).first()
    return asesor if asesor else None

def dias_de_mora(instance):
    hoy = date.today()

    fecha = instance.get('fecha_entrar_en_mora')

    if instance.get('is_paid_off',''):
        return 0

    if not fecha:
        return 0
    

    
    if isinstance(fecha, str):
        try:
            fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
        except ValueError:
            return 0  # fecha inválida

    
    if isinstance(fecha, datetime):
        fecha = fecha.date()

    
    if hoy <= fecha:
        return 0

    return (hoy - fecha).days

def get_creation_date(credito):
    fecha = credito.get('credito', {}).get('creation_date')

    if not fecha:
        return datetime.min  # Para que quede al inicio

    if isinstance(fecha, str):
        try:
            return datetime.strptime(fecha, '%Y-%m-%d')
        except ValueError:
            return datetime.min

    if isinstance(fecha, date):
        return datetime.combine(fecha, datetime.min.time())

    if isinstance(fecha, datetime):
        return fecha

    return datetime.min


def crear_excel_creditos(datos, dia = None):
    if dia is None:
        dia = datetime.now().date()

    workbook = Workbook()
    workbook.remove(workbook.active)

    filtros = {
        'Creditos Nuevos': creditos_creado_ese_dia(datos,dia),
        'Todos Los Creditos': datos,
        'Creditos Cancelados': creditos_cancelados(datos),
        'Creditos Estado Juridico':creditos_estado_juridico(datos),
        'Creditos con Excedente':creditos_con_excedentes(datos),
        'Creditos en Atraso':creditos_en_atraso(datos),
        'Creditos con falta de Aportacion':creditos_falta_aportacion(datos),
    }

    # Agregar encabezado    
    encabezados = [
        "#", "FECHA DE REGISTRO", "CODIGO DEL CREDITO", 
        "CLIENTE", "MONTO OTORGADO", "PROPOSITO", "PLAZO EN MESES", "TASA DE INTERES",
        "FORMA DE PAGO", "TIPO DE CREDITO", "DESEMBOLSO","FECHA DE INICIO DEL CREDITO", 
        "FECHA DE VENCIMIENTO DEL CREDITO", "FECHA LIMITE DE PAGO", 
        "SALDO ACTUAL", "SALDO CAPITAL PENDIENTE","SALDO EXCEDENTE" ,"DIAS DE MORA","STATUS POR FECHAS","STATUS POR APORTACION","STATUS JUDICIAL","STATUS CANCELADO", 
        "NUMERO DE REFERENCIA", "ASESOR DE CREDITO"
    ]

    for filtro,data in filtros.items():
        sheet = workbook.create_sheet(title=filtro[:31])

        data = sorted(data, key=get_creation_date)

        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Agregar los datos
        contador = 0
        for idx, credito in enumerate(data, start=2):
            contador+=1
            informacion_credito = credito.get('credito', {})
            cuota = credito.get('cuota_vigente')
            fecha_limite = formater_fecha(cuota.get('fecha_limite')) if cuota else ''
            numero_referencia = cuota.get('numero_referencia') if cuota else ''

            # Mensajes de estado
            if credito['credito']['estado_aportacion']:
                mensaje = 'VIGENTE'
            elif credito['credito']['estado_aportacion'] is None:
                mensaje = 'SIN APORTACIONES'
            else:
                mensaje = 'EN ATRASO'

            aportacion = mensaje
            estado_fechas = 'VIGENTE' if informacion_credito.get('estados_fechas','') else 'EN ATRASO'

            saldo_actual = informacion_credito.get('saldo_actual','')
            saldo_capital_pendiente = informacion_credito.get('saldo_pendiente','')
            saldo_excedente = informacion_credito.get('excedente','')

            es_credito_judicial = 'NO'
            es_credito_cancelado = 'NO'

            if informacion_credito.get('is_paid_off',''):
                es_credito_cancelado = 'SI'
                es_credito_judicial = 'NO'
                aportacion = 'VIGENTE'
                estado_fechas = 'VIGENTE'
                
            
            if informacion_credito.get('estado_judicial',''):
                es_credito_cancelado = 'NO'
                es_credito_judicial = 'SI'

            
            if  saldo_actual < 0:
                saldo_excedente = abs(informacion_credito.get('saldo_actual',''))
                saldo_actual = 0
                saldo_capital_pendiente = 0
            
            if  saldo_capital_pendiente < 0:
                saldo_excedente = abs(informacion_credito.get('saldo_pendiente',''))
                saldo_actual = 0
                saldo_capital_pendiente = 0
            
            if  saldo_excedente < 0:
                saldo_excedente = abs(informacion_credito.get('excedente',''))
                saldo_actual = 0
                saldo_capital_pendiente = 0



            fila = [
                str(contador),
                f"{informacion_credito.get('creation_date', '')}",
                f"{informacion_credito.get('codigo_credito', '')}",
                f"{get_cliente(informacion_credito.get('customer_id_id', ''))}",
                f"{formatear_numero(informacion_credito.get('monto', 0))}",
                f"{informacion_credito.get('proposito', '')}",

                f"{informacion_credito.get('plazo', '')}",
                f"{informacion_credito.get('tasa_interes', 0) * 100}%",
                f"{informacion_credito.get('forma_de_pago', '')}",
                f"{informacion_credito.get('tipo_credito', '')}",

                # Desembolsos - manejo seguro de lista
                f"{credito.get('desembolsos', [{}])[0].get('forma_desembolso', '')}",
                f"{informacion_credito.get('fecha_inicio', '')}",
                f"{informacion_credito.get('fecha_vencimiento', '')}",

                f"{fecha_limite}",

                f"{formatear_numero(saldo_actual)}",
                f"{formatear_numero(saldo_capital_pendiente)}",
                f"{formatear_numero(saldo_excedente)}",
                str(dias_de_mora(informacion_credito)),
                

                f"{estado_fechas}",
                str(aportacion),
                str(es_credito_judicial),
                str(es_credito_cancelado),

                f"{numero_referencia}",
                f"{get_asesor_credito(informacion_credito.get('asesor_de_credito_id', ''))}",

            ]

            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)


    return workbook

