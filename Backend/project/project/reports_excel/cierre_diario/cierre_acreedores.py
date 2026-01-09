# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile

# Tiempo
from datetime import datetime, timedelta

# Modelo
from apps.customers.models import Customer, CreditCounselor
from apps.financings.formato import formatear_numero

# Filtros
from .filtros.acreedores import *
from .funcion import safe_list_get



def crear_excel_creditos_asesores(datos, dia = None):
    if dia is None:
        dia = datetime.now().date()

    workbook = Workbook()
    workbook.remove(workbook.active)

    filtros = {
        'Acreedores Nuevos': creditos_creado_ese_dia(datos,dia),
        'Todos Los Acreedores': datos,
        'Acreedores Cancelados': creditos_cancelados(datos),
        'Acreedores con Excedente':creditos_con_excedentes(datos),
        'Acreedores en Atraso':creditos_en_atraso(datos),
        'Acreedores con falta de Aportacion':creditos_falta_aportacion(datos),
    }

    # Agregar encabezado    
    encabezados = [
        "#", "FECHA DE REGISTRO","CODIGO DE ACREEDOR", "NOMBRE DEL ACREEDOR",
        "MONTO OTORGADO","PROPOSITO","PLAZO","TASA","FECHA DE INICIO","FECHA DE VENCIMIENTO","FECHA LIMITE",
        "SALDO ACTUAL","SALDO CAPITAL PENDIENTE","SALDO EXCEDENTE","STATUS POR FECHAS","STATUS POR APORTACION","STATUS CANCELADO",
        "NUMERO DE REFERENCIA"
    ]

    for filtro,data in filtros.items():
        sheet = workbook.create_sheet(title=filtro[:31])

        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Agregar los datos
        contador = 0
        for idx, credito in enumerate(data, start=2):
            contador+=1
            informacion_credito = credito.get('informacion_acreedor', {})

            # Mensajes de estado
            if credito['informacion_acreedor']['estado_aportacion']:
                mensaje = 'VIGENTE'
            elif credito['informacion_acreedor']['estado_aportacion'] is None:
                mensaje = 'SIN APORTACIONES'
            else:
                mensaje = 'EN ATRASO'

            
            aportacion = mensaje
            estado_fechas = 'VIGENTE' if informacion_credito.get('estados_fechas','') else 'EN ATRASO'

            saldo_actual = informacion_credito.get('saldo_actual','')
            saldo_capital_pendiente = informacion_credito.get('saldo_pendiente','')
            saldo_excedente = informacion_credito.get('excedente','')

            
            es_credito_cancelado = 'NO'

            if informacion_credito.get('is_paid_off',''):
                es_credito_cancelado = 'SI'                
                aportacion = 'VIGENTE'
                estado_fechas = 'VIGENTE'

            if  saldo_actual < 0:
                saldo_excedente = abs(informacion_credito.get('saldo_actual',''))
                saldo_actual = 0
                saldo_capital_pendiente = 0
            
            if  saldo_capital_pendiente < 0:
                saldo_excedente = abs(informacion_credito.get('saldo_pendiente',''))
                saldo_actual = 0
                saldo_capital_pendiente = 0
            
            if  saldo_excedente < 0:
                saldo_excedente = abs(informacion_credito.get('excedente',''))
                saldo_actual = 0
                saldo_capital_pendiente = 0

            cuota = credito.get('cuota_vigente')
            fecha_limite = formater_fecha(cuota.get('fecha_limite')) if cuota else ''
            numero_referencia = credito.get('informacion_acreedor', {}).get('numero_referencia', '')

            fila = [
                str(contador),
                f"{credito.get('informacion_acreedor', {}).get('creation_date', '')}",
                f"{credito.get('informacion_acreedor', {}).get('codigo_acreedor', '')}",
                f"{credito.get('informacion_acreedor', {}).get('nombre_acreedor', '')}",
                f"{formatear_numero(credito.get('informacion_acreedor', {}).get('monto', 0))}",
                f"{credito.get('informacion_acreedor', {}).get('observaciones', '')}",

                f"{credito.get('informacion_acreedor', {}).get('plazo', '')}",
                f"{credito.get('informacion_acreedor', {}).get('tasa', 0) * 100}%",
                

                # Desembolsos - manejo seguro de lista
                
                f"{credito.get('informacion_acreedor', {}).get('fecha_inicio', '')}",
                f"{credito.get('informacion_acreedor', {}).get('fecha_vencimiento', '')}",
                f"{fecha_limite}",

                f"{formatear_numero(saldo_actual)}",
                f"{formatear_numero(saldo_capital_pendiente)}",
                f"{formatear_numero(saldo_excedente)}",
                
                # Estados
                str(estado_fechas),
                str(aportacion),
                str(es_credito_cancelado),

                f"{numero_referencia}",
               

            ]

            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)


    return workbook

