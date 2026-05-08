# Modelos
from apps.actividades.models import InformeDiarioSistema, DetalleInformeDiario

# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile


from django.db.models import Q, Sum
# Funciones
from .cierre_clientes import crear_excel_clientes
from .cierre_creditos import crear_excel_creditos
from .cierre_acreedores import crear_excel_creditos_asesores
from .cierre_seguros import crear_excel_creditos_seguros
from .cierre_ingresos import crear_excel_ingresos
from .cierre_egresos import crear_excel_egresos
from .cierre_bancos import crear_excel_bancos
from .cierre_pagos import crear_excel_pagos
from .cierre_recibos import crear_excel_recibos
from .cierre_facturas import crear_excel_facturas
# Tiempo
from datetime import datetime

# template
from django.views.generic import TemplateView
from django.http import HttpResponse


# Decoradores
from project.decorador import permiso_requerido
from django.utils.decorators import method_decorator

class CierreDiario(TemplateView):
    
    def get_queryset(self):
        try:
            sucursal = getattr(self.request,'sucursal_actual',None)
            # Asignar la consulta a una variable local
            query = self.query()

            # Crear una lista para almacenar los filtros
            filters = Q()

            
         

            qs = InformeDiarioSistema.objects.filter(sucursal=sucursal)

            if query:
                filters = Q()

                # 🔹 Opción 1: buscar por ID
                if query.isdigit():
                    filters = Q(id=int(query))

                else:
                    # 🔹 Opción 2: buscar por fecha (YYYY-MM-DD)
                    try:
                        fecha = datetime.strptime(query, '%Y-%m-%d')
                        fecha_inicio = fecha.replace(hour=0, minute=0, second=0)
                        fecha_fin = fecha.replace(hour=23, minute=59, second=59)
                        filters = Q(fecha_registro__range=(fecha_inicio, fecha_fin))
                    except ValueError:
                        pass

                qs = qs.filter(filters)

            return qs.order_by('-id').first()
        
        except Exception as e:
            # Manejar cualquier excepción que ocurra y devolver un queryset vacío
            print(f"Error al filtrar el queryset: {e}")
            return InformeDiarioSistema.objects.filter(sucursal=sucursal).order_by('-id').first()
        
    def query(self):
        return self.request.GET.get('q', '').strip()

    
    def get(self, request, *args, **kwargs):
        # Crear el archivo Excel
        zip_buffer = io.BytesIO()
        zip_file = zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED)

        # Obtener datos
        informe = self.get_queryset()
        

        if informe is None:
            return HttpResponse("No se encontró información del informe", status=404)
        
        dia =  informe.fecha_registro

        workbook_main = Workbook()
        sheet = workbook_main.active
        sheet.title = "Reporte de Cierre Diario"

        # Encabezados
        encabezados = [
            "#", "Nombre", "Total"
        ]
        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        

        registros = DetalleInformeDiario.objects.filter(reporte__id=informe.id)

        contador = 0

        for idx, registro in enumerate(registros, start=2):
            contador += 1
            

            fila = [contador,str( registro.tipo_datos),str(registro.cantidad)]

            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)

            # Creacion individual del reporte
            cierre_diario_buffer = io.BytesIO()

            if registro.tipo_datos == 'clientes' and registro.data['clientes'] is not None:
                wb_cliente = crear_excel_clientes(registro.data['clientes'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_clientes.xlsx", cierre_diario_buffer.read())
            
            if registro.tipo_datos == 'creditos' and registro.data['creditos'] is not None:
                wb_cliente = crear_excel_creditos(registro.data['creditos'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_creditos.xlsx", cierre_diario_buffer.read())
            
            if registro.tipo_datos == 'acreedores' and registro.data['acreedores'] is not None:
                wb_cliente = crear_excel_creditos_asesores(registro.data['acreedores'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_acreedores.xlsx", cierre_diario_buffer.read())

            if registro.tipo_datos == 'seguros' and registro.data['seguros'] is not None:
                wb_cliente = crear_excel_creditos_seguros(registro.data['seguros'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_seguros.xlsx", cierre_diario_buffer.read())
            
            if registro.tipo_datos == 'ingresos' and registro.data['ingresos'] is not None:
                wb_cliente = crear_excel_ingresos(registro.data['ingresos'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_ingresos.xlsx", cierre_diario_buffer.read())
            
            if registro.tipo_datos == 'egresos' and registro.data['egresos'] is not None:
                wb_cliente = crear_excel_egresos(registro.data['egresos'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_egresos.xlsx", cierre_diario_buffer.read())
            
            if registro.tipo_datos == 'bancos' and registro.data['bancos'] is not None:
                wb_cliente = crear_excel_bancos(registro.data['bancos'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_bancos.xlsx", cierre_diario_buffer.read())
            
            if registro.tipo_datos == 'pagos' and registro.data['pagos'] is not None:
                wb_cliente = crear_excel_pagos(registro.data['pagos'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_pagos.xlsx", cierre_diario_buffer.read())
            
            if registro.tipo_datos == 'recibos' and registro.data['recibos'] is not None:
                wb_cliente = crear_excel_recibos(registro.data['recibos'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_recibos.xlsx", cierre_diario_buffer.read())

            if registro.tipo_datos == 'facturas' and registro.data['facturas'] is not None:
                wb_cliente = crear_excel_facturas(registro.data['facturas'], dia)
                wb_cliente.save(cierre_diario_buffer)
                cierre_diario_buffer.seek(0)
                zip_file.writestr(f"reporte_facturas.xlsx", cierre_diario_buffer.read())

        # Crear respuesta de descarga

        # 3. Guardar el Excel general en el zip
        main_excel_buffer = io.BytesIO()
        workbook_main.save(main_excel_buffer)
        main_excel_buffer.seek(0)
        zip_file.writestr("reporte_cierre_diario.xlsx", main_excel_buffer.read())

        zip_file.close()

        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_de_cierre_diario_{dia}.zip"'
        return response

