# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile

# Tiempo
from datetime import datetime, timedelta

# Funciones
from apps.financings.formato import formatear_numero




def crear_excel_pagos(data, dia = None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informacion Pagos"

    encabezados = [
        "#","FECHA","NUMERO DE REFERENCIA", "MONTO", "DESCRIPCION", "TIPO DE PAGO",
        "ESTADO DE TRANSACCION","REGISTRO FICTICIO"
    ]


    for col_idx, header in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

 

    contador = 0  # Para la numeración real del archivo

    for idx, pago in enumerate(data, start=2):
        
        contador+=1
        informacion_pago = pago.get('informacion_pagos',{})
      
        fila = [
            contador, 
            str(informacion_pago.get('fecha_emision','')),
            informacion_pago.get('numero_referencia',''),
            formatear_numero(informacion_pago.get('monto','')),
            informacion_pago.get('descripcion',''),
            informacion_pago.get('tipo_pago',''),
            informacion_pago.get('estado_transaccion',''),
            informacion_pago.get('registro_ficticio','')
            
            ]

        for col_idx, value in enumerate(fila, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)


    return workbook
