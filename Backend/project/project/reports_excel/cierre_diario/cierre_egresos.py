# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile

# Tiempo
from datetime import datetime, timedelta

# Funciones
from apps.financings.formato import formatear_numero




def crear_excel_egresos(data, dia = None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informacion Egresos"

    encabezados = [
        "#","CODIGO DE EGRESO","DESCRIPCION",
        "OBSERVACIONES","NUMERO DE REFERENCIA", "FECHA","FECHA DE DOCUMENTO FISCAL",
        "NUMERO DE DOCUMENTO FISCAL","NIT","MONTO","MONTO DE DOCUMENTO","NOMBRE DE COLABORADOR",
        "PAGO CORRESPONDIENTE","TIPO DE IMPUESTO", "TIPO DE GASTO"

    ]


    for col_idx, header in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

 

    contador = 0  # Para la numeración real del archivo

    for idx, egreso in enumerate(data, start=2):
        
        contador+=1
        informacion_egreso = egreso.get('informacion_egreso',{})
      
       

        fila = [
            contador, 
            informacion_egreso.get('codigo_egreso',''),
            informacion_egreso.get('descripcion',''),
            informacion_egreso.get('observaciones',''),
            informacion_egreso.get('numero_referencia',''),
            str(informacion_egreso.get('fecha','')),
            str(informacion_egreso.get('fecha_doc_fiscal','')),
            informacion_egreso.get('numero_doc',''),
            informacion_egreso.get('nit',''),
            str(formatear_numero(informacion_egreso.get('monto',''))),
            str(formatear_numero(informacion_egreso.get('monto_doc',''))),
            informacion_egreso.get('nombre',''),
            informacion_egreso.get('pago_correspondiente',''),
            informacion_egreso.get('tipo_impuesto',''),
            informacion_egreso.get('tipo_gasto',''),
            ]

        for col_idx, value in enumerate(fila, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)


    return workbook
