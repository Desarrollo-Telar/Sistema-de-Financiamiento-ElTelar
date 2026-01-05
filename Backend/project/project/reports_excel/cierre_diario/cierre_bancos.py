# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile

# Tiempo
from datetime import datetime, timedelta

# Funciones
from apps.financings.formato import formatear_numero




def crear_excel_bancos(data, dia = None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informacion Bancos"

    encabezados = [
        "#","FECHA","NUMERO DE REFERENCIA","SECUENCIAL","CHEQUE",
        "CREDITO","DEBITO","DESCRIPCION","SALDO CONTABLE",
        "SALDO DISPONIBLE","REGISTRO FICTICIO"

    ]


    for col_idx, header in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

 

    contador = 0  # Para la numeración real del archivo

    for idx, banco in enumerate(data, start=2):
        
        contador+=1
        informacion_banco = banco.get('informacion_banco',{})
      
       

        fila = [
            contador, 
            str(informacion_banco.get('fecha','')),
            informacion_banco.get('referencia',''),
            informacion_banco.get('secuencial',''),
            informacion_banco.get('cheque',''),
            informacion_banco.get('credito',''),
            informacion_banco.get('debito',''),
            informacion_banco.get('descripcion',''),
            formatear_numero(informacion_banco.get('saldo_contable','')),
            formatear_numero(informacion_banco.get('saldo_disponible','')),
            informacion_banco.get('registro_ficticio',''),
            ]

        for col_idx, value in enumerate(fila, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)


    return workbook
