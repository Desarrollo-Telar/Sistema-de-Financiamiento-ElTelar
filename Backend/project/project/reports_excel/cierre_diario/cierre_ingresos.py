# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile

# Tiempo
from datetime import datetime, timedelta

# Funciones
from apps.financings.formato import formatear_numero




def crear_excel_ingresos(data, dia = None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informacion Ingresos"

    encabezados = [
        "#","CODIGO DE INGRESO","DESCRIPCION",
        "OBSERVACIONES","NUMERO DE REFERENCIA", "FECHA","MONTO"

    ]


    for col_idx, header in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

 

    contador = 0  # Para la numeración real del archivo

    for idx, ingreso in enumerate(data, start=2):
        
        contador+=1
        informacion_ingreso = ingreso.get('informacion_ingreso',{})
      
       

        fila = [
            contador, 
            informacion_ingreso.get('codigo_ingreso',''),
            informacion_ingreso.get('descripcion',''),
            informacion_ingreso.get('observaciones',''),
            informacion_ingreso.get('numero_referencia',''),
            str(informacion_ingreso.get('fecha','')),
            str(formatear_numero(informacion_ingreso.get('monto','')))
          
            ]

        for col_idx, value in enumerate(fila, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)


    return workbook
