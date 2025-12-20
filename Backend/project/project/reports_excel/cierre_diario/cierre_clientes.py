# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile

# Tiempo
from datetime import datetime, timedelta
from .filtros.creditos import formater_fecha
from .funcion import safe_list_get

def get_puesto(source_of_income, cliente):
        puesto_cliente = ''
        informacion_labora = safe_list_get(cliente.get('informacion_laboral'))

        if informacion_labora is not None:

            if source_of_income != 'Otra':
                puesto_cliente = informacion_labora.get('position','')
            else:
                puesto_cliente = 'Otra Fuente de Ingreso'

            return puesto_cliente

def get_empresa_laburo(source_of_income, cliente):
        empresa_laburo_cliente = ''
        informacion_labora = safe_list_get(cliente.get('informacion_laboral'))

        if informacion_labora is not None:
            if source_of_income != 'Otra':
                empresa_laburo_cliente = informacion_labora.get('company_name','')
            else:
                
                empresa_laburo_cliente = source_of_income

        return empresa_laburo_cliente

def get_estado_laboral(source_of_income, cliente):
        estado_laboral_cliente = ''
        informacion_labora = safe_list_get(cliente.get('informacion_laboral'))

        if informacion_labora is not None:

            if source_of_income != 'Otra':
                estado_laboral_cliente = informacion_labora.get('employment_status','')
            else:
                estado_laboral_cliente = 'COMPLETO'

        return estado_laboral_cliente

def get_edad(date_birth):
        from datetime import date
        """
        Calcula la edad de la persona en base a su fecha de nacimiento.
        Devuelve None si la fecha de nacimiento no está definida.
        """
        if not date_birth:
            return None  # O 0, según cómo quieras manejarlo

        today = date.today()
        date_birth = formater_fecha(date_birth)
        age = today.year - date_birth.year

        # Si aún no ha llegado su cumpleaños este año, restamos 1
        if (today.month, today.day) < (date_birth.month, date_birth.day):
            age -= 1

        return age

def crear_excel_clientes(data, dia = None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informacion Personal"

    encabezados = [
        "#",
        "NOMBRE",
        "TIPO DE IDENTIFICACION",
        "NUMERO DE IDENTIFICACION",
        "NUMERO DE TELEFONO",
        "EDAD",
        "GENERO",
        "CODIGO DE CLIENTE",
        "PROFESION U OFICIO",
        "ASESOR DEL CREDITO",
        "DIRECCION DEL CLIENTE",
        "MUNICIPIO DEL CLIENTE",
        "DEPARTAMENTO DEL CLIENTE",
        "DIRECCION DE TRABAJO",
        "MUNICIPIO DE LABORAL",
        "DEPARTAMENTO DE LABORAL",
        "FUENTE DE INGRESO",
        "ESTADO LABORAL",
        "EMPRESA DE LABORAL",
        "PUESTO",
        "REFERENCIA 1",
        "REFERENCIA 2",
        "REFERENCIA 3",
        "REFERENCIA 4",
    ]


    for col_idx, header in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

 

    contador = 0  # Para la numeración real del archivo

    for idx, cliente in enumerate(data, start=2):
        
        contador+=1
        # Direcciones (manejo seguro de listas)
        direccion_0 = safe_list_get(cliente.get('direcciones'), 0, {})
        direccion_1 = safe_list_get(cliente.get('direcciones'), 1, {})
        
        informacion_labora = safe_list_get(cliente.get('informacion_laboral'), 0, {})

        informacion_referencias_0 = safe_list_get(cliente.get('informacion_referencias'), 0, {})
        informacion_referencias_1 = safe_list_get(cliente.get('informacion_referencias'), 1, {})
        informacion_referencias_2 = safe_list_get(cliente.get('informacion_referencias'), 2, {})
        informacion_referencias_3 = safe_list_get(cliente.get('informacion_referencias'), 3, {})
       

        fila = [
            contador, 
            # Información personal
            f"{cliente.get('informacion_personal', {}).get('first_name', '')} {cliente.get('informacion_personal', {}).get('last_name', '')}",
            f"{cliente.get('informacion_personal', {}).get('type_identification', '')}",
            f"{cliente.get('informacion_personal', {}).get('identification_number', '')}",
            f"{cliente.get('informacion_personal', {}).get('telephone', '')}",
            f"{str(get_edad(cliente.get('informacion_personal', {}).get('date_birth', '')))}",
            f"{cliente.get('informacion_personal', {}).get('gender', '')}",
            f"{cliente.get('informacion_personal', {}).get('customer_code', '')}",
            f"{cliente.get('informacion_personal', {}).get('profession_trade', '')}",
            f"{cliente.get('informacion_personal', {}).get('asesor', '')}",
            


            

            f"{direccion_0.get('street', '')}",
            f"{direccion_0.get('state', '')}",
            f"{direccion_0.get('city', '')}",
            f"{direccion_1.get('street', '')}",
            f"{direccion_1.get('state', '')}",
            f"{direccion_1.get('city', '')}",
            # Información laboral
            
            f"{informacion_labora.get('source_of_income', '')}",
            f"{get_estado_laboral(informacion_labora.get('source_of_income', ''), cliente)}",
            f"{get_empresa_laburo(informacion_labora.get('source_of_income', ''), cliente)}",
            f"{get_puesto(informacion_labora.get('source_of_income', ''), cliente)}",
            
            
            
            # Referencias (manejo seguro de listas)
            f"{informacion_referencias_0.get('full_name', '')}",
            f"{informacion_referencias_1.get('full_name', '')}",
            f"{informacion_referencias_2.get('full_name', '')}",
            f"{informacion_referencias_3.get('full_name', '')}",
            ]

        for col_idx, value in enumerate(fila, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)


    return workbook