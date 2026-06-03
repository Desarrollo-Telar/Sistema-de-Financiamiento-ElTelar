# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile

# Tiempo
from datetime import datetime, timedelta

# Funciones
from apps.financings.formato import formatear_numero

# Modelo
from apps.financings.models import Payment


def obtener_mapeo_pagos(data):
    """Extrae todos los pago_id y busca los pagos en una sola consulta SQL"""
    pago_ids = []
    for recibo in data:
        info = recibo.get('informacion_recibo', {})
        pago_id = info.get('pago_id')
        if pago_id:
            pago_ids.append(pago_id)
            
    if not pago_ids:
        return {}

    # Traemos todos los pagos y sus relaciones en UN SOLO viaje a la base de datos
    pagos = Payment.objects.filter(id__in=pago_ids).select_related(
        'acreedor', 'seguro', 'credit', 'cliente'
    )
    
    # Lo convertimos a un diccionario para buscar por ID en tiempo récord O(1)
    return {str(pago.id): pago for pago in pagos}


def recibo_para_optimizado(pago_id, mapeo_pagos):
    mensaje = ''
    if not pago_id:
        return mensaje
        
    # Buscamos en el diccionario en memoria, no en la base de datos
    pago = mapeo_pagos.get(str(pago_id))
    if not pago:
        return mensaje
       
    if pago.acreedor is not None:
        mensaje = f'{pago.acreedor.nombre_acreedor}'
    elif pago.seguro is not None:
        mensaje = f'{pago.seguro.nombre_acreedor}'
    elif pago.credit is not None:
        mensaje = f'{pago.credit.customer_id}'
    elif pago.cliente is not None:
        mensaje = f'{pago.cliente}'

    return mensaje


def crear_excel_recibos(data, dia = None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informacion Recibos"

    encabezados = [
        "#","FECHA","NUMERO DE RECIBO", "PARA", "INTERES", "INTERES PAGAGO",
        "MORA","MORA PAGADA","CAPITAL APORTADO","TOTAL","ESTA FACTURADO"
    ]

    for col_idx, header in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # --- AQUÍ LA MAGIA ---
    # Generamos el mapa de pagos ANTES de empezar el bucle
    mapeo_pagos = obtener_mapeo_pagos(data)

    contador = 0

    for idx, recibo in enumerate(data, start=2):
        contador += 1
        informacion_recibo = recibo.get('informacion_recibo', {})
        pago_id = informacion_recibo.get('pago_id', '')
        
        fila = [
            contador, 
            str(informacion_recibo.get('fecha', '')),
            informacion_recibo.get('recibo', ''),
            # Pasamos el ID y el mapa ya cargado en memoria
            str(recibo_para_optimizado(pago_id, mapeo_pagos)),
            formatear_numero(informacion_recibo.get('interes', '')),
            formatear_numero(informacion_recibo.get('interes_pagado', '')),
            formatear_numero(informacion_recibo.get('mora', '')),
            formatear_numero(informacion_recibo.get('mora_pagada', '')),
            formatear_numero(informacion_recibo.get('aporte_capital', '')),
            formatear_numero(informacion_recibo.get('total', '')),
            informacion_recibo.get('factura', '')
        ]

        for col_idx, value in enumerate(fila, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)

    return workbook