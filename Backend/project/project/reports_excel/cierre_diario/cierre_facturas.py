# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json
import io
import zipfile

# Tiempo
from datetime import datetime, timedelta

# Funciones
from apps.financings.formato import formatear_numero

# Modelo
from apps.financings.models import Recibo

def informacion_recibo(instance):

    if instance is None:
        return ''
    
    recibo = Recibo.objects.get(id=instance)

    if recibo is None:
        return ''
    
    return formatear_numero(recibo.pago.monto)
    

def crear_excel_facturas(data, dia = None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informacion Facturas"

    encabezados = [
        "#","FECHA","NUMERO DE FACTURA", "NOMBRE DEL RECEPTOR","NIT",
        "NUMERO DE AUTORIZACION", "SERIE DE AUTORIZACION", "IDENTIFICADOR","MONTO"
    ]


    for col_idx, header in enumerate(encabezados, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

 

    contador = 0  # Para la numeración real del archivo

    for idx, factura in enumerate(data, start=2):
        
        contador+=1
        informacion_factura = factura.get('informacion_factura',{})
        
        
      
        fila = [
            contador, 
            str(informacion_factura.get('issue_date','')),
            informacion_factura.get('numero_factura',''),
            informacion_factura.get('nombre_receptor',''),
            informacion_factura.get('nit_receptor',''),
            informacion_factura.get('numero_autorizacion',''),
            informacion_factura.get('serie_autorizacion',''),
            informacion_factura.get('identificador',''),
            str(informacion_recibo(informacion_factura.get('recibo_id_id','')))
            
            ]

        for col_idx, value in enumerate(fila, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)


    return workbook
