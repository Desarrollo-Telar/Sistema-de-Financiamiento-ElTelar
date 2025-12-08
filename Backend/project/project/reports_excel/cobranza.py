from openpyxl import Workbook
from django.http import HttpResponse
import json

from django.db.models import Q, Sum
from datetime import datetime, timedelta


# Modelos
from apps.customers.models import HistorialCobranza, CreditCounselor
from apps.users.models import User

from django.views.generic import TemplateView
from django.shortcuts import render, get_object_or_404, redirect


class ReporteCobranza(TemplateView):
    def get_queryset(self):
        try:
            request = self.request

            query = self.query()
            usuario = self.usuario()
            mes = self.query_mes()
            anio = self.query_anio()

            

            # Crear una lista para almacenar los filtros
            filters = Q()
            
            if usuario:
                usuario = User.objects.get(id=usuario)
                filters &= Q(usuario=usuario)
            
            if mes:
                filters &= Q(fecha_cambio__month=mes)
            
            if anio:
                filters &= Q(fecha_cambio__year=anio)


            
            return HistorialCobranza.objects.filter(filters).order_by('-id')
        
        except Exception as e:
            # Manejar cualquier excepción que ocurra y devolver un queryset vacío
            print(f"Error al filtrar el queryset: {e}")
            return HistorialCobranza.objects.none()

    def query(self):
        return self.request.GET.get('q')
    
    def usuario(self):
        return self.request.GET.get('usuario')
    
    def query_mes(self):
        return self.request.GET.get('mes')
    
    def query_anio(self):
        return self.request.GET.get('anio')
    
    def get(self, request, *args, **kwargs):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Cartera de Cobranza"
        encabezados = [
            "#", "RESPONSABLE DE LA GESTION", "ACCION", "FECHA DE REGISTRO", "CODIGO DEL CREDITO", "CLIENTE", "TIPO DE COBRANZA", 
            "TIPO DE GESTION", "RESULTADO","ESTADO DE COBRANZA","FECHA DE SEGUIMIENTO O PROMESA DE PAGO","OBSERVACIONES"
        ]

        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Obtener datos
        informacion_consultada = self.get_queryset()
        contador = 0

        for idx, reporte in enumerate(informacion_consultada, start=2):
            contador += 1
            cobranza = reporte.cobranza
            

            fila = [
                contador,
                str(reporte.usuario),
                reporte.accion,
                str(reporte.fecha_cambio.date()),
                str(cobranza.credito.codigo_credito),
                str(cobranza.credito.customer_id),
                cobranza.tipo_cobranza,
                cobranza.tipo_gestion,
                cobranza.resultado,
                cobranza.estado_cobranza,                
                str(cobranza.get_fecha_seg_o_promesa()),
                cobranza.observaciones,      
            ]

            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)

        # Crear respuesta de descarga
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_cobranza_{timestamp}.xlsx"'
        workbook.save(response)
        return response
    