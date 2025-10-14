
# Modelos
from apps.accountings.models import Egress, Income

# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json


from django.db.models import Q, Sum

# Tiempo
from datetime import datetime

# template
from django.views.generic import TemplateView
from django.http import HttpResponse


# Decoradores
from project.decorador import permiso_requerido
from django.utils.decorators import method_decorator


class ReporteEgresosExcelView(TemplateView):

    
    def get_queryset(self):
        try:
            sucursal = self.request.session['sucursal_id']
            query = self.query()
            mes = self.mes_reporte()
            anio = self.anio_reporte()

            # Crear una lista para almacenar los filtros
            filters = Q()

            # Añadir filtros si la consulta no está vacía
            if query:
                filters |= Q(fecha__icontains=query)
                filters |= Q(fecha_doc_fiscal__icontains=query)                
                filters |= Q(numero_doc__icontains=query)                
                filters |= Q(codigo_egreso__icontains=query)
                filters |= Q(numero_referencia__icontains=query)
                filters |= Q(nit__icontains=query)
                filters |= Q(nombre__icontains=query)

                # Si la consulta es numérica, usar filtro exacto para campos numéricos
                if query.isdigit():
                    filters |= Q(monto__exact=query)
                    
            if mes:
                filters &= Q(fecha__month=mes)

            if anio:
                filters &= Q(fecha__year=anio)

            # Filtrar los objetos Banco usando los filtros definidos
            return Egress.objects.filter(filters, sucursal=sucursal).order_by('-id')
        except Exception as e:
            # Manejar cualquier excepción que ocurra y devolver un queryset vacío
            print(f"Error al filtrar el queryset: {e}")
            return Egress.objects.none()
    

    def query(self):
        return self.request.GET.get('q')
    
    def mes_reporte(self):
        return self.request.GET.get('mes')
    
    def anio_reporte(self):
        return self.request.GET.get('anio')


    def get(self, request, *args, **kwargs):
        # Crear el archivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de EGRESOS"

        # Encabezados
        encabezados = [
            "#", "CODIGO DE EGRESOS", "FECHA", "FECHA DE DOCUMENTO FISCAL", "NUMERO DE DOCUMENTO FISCAL", "NIT", "MONTO","MONTO DEL DOCUMENTO" ,"NUMERO DE REFERENCIA", "DESCRIPCION", "OBSERVACIONES",
            "NOMBRE DEL COLABORADOR", "PAGO CORRESPONDIENTE", "TIPO DE IMPUESTO", "TIPO DE GASTO",
        ]
        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Obtener datos
        informacion_consultada = self.get_queryset()

        contador = 0

        for idx, egreso in enumerate(informacion_consultada, start=2):
            contador += 1
            
            fila = [
                contador,
                egreso.codigo_egreso,
                egreso.fecha,
                egreso.fecha_doc_fiscal,
                egreso.numero_doc,
                egreso.nit,
                egreso.monto,
                egreso.monto_doc,
                egreso.numero_referencia,
                egreso.descripcion,
                egreso.observaciones,
                egreso.nombre,
                egreso.pago_correspondiente,
                egreso.tipo_impuesto,
                egreso.tipo_gasto
            ]

            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)

        # Crear respuesta de descarga
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_egresos_{timestamp}.xlsx"'
        workbook.save(response)
        return response

