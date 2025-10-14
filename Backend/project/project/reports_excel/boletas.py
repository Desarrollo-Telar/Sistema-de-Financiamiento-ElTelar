# Modelos
from apps.financings.models import Payment, Recibo
from apps.accountings.models import Egress, Income

# Reporte Excel
from openpyxl import Workbook
from django.http import HttpResponse
import json


from django.db.models import Q, Sum

# Tiempo
from datetime import datetime

# template
from django.views.generic import TemplateView
from django.http import HttpResponse


# Decoradores
from project.decorador import permiso_requerido
from django.utils.decorators import method_decorator



class ReporteBaseBoletasExcelView(TemplateView):

    def get_queryset(self):
        try:
            sucursal = self.request.session['sucursal_id']
            # Asignar la consulta a una variable local
            query = self.query()

            # Crear una lista para almacenar los filtros
            filters = Q()

            # Añadir filtros si la consulta no está vacía
            if query:
                try:
                    fecha = datetime.strptime(query, '%Y-%m-%d')
                    fecha_inicio = datetime.combine(fecha.date(), datetime.min.time())
                    fecha_fin = datetime.combine(fecha.date(), datetime.max.time())
                    filters |= Q(fecha_emision__range=(fecha_inicio, fecha_fin))
                except ValueError:
                    pass  # No es fecha válida, continúa con los otros filtros


            # Definir los filtros utilizando Q objects
            filters |= Q(fecha_emision__icontains=query)
            filters |= Q(numero_referencia__icontains=query)
            filters |= Q(estado_transaccion__icontains=query)
            filters |= Q(credit__codigo_credito__icontains=query)
            filters |= Q(tipo_pago__icontains=query)

 
            # Filtrar los objetos Customer usando los filtros definidos
            return Payment.objects.filter(filters, sucursal=sucursal)
        except Exception as e:
            # Manejar cualquier excepción que ocurra y devolver un queryset vacío
            print(f"Error al filtrar el queryset: {e}")
            return Payment.objects.filter(
                registro_ficticio=False,
                estado_transaccion='COMPLETADO'
            ).order_by('-id')
        
    def query(self):
        return self.request.GET.get('q')
    
    @method_decorator(permiso_requerido('puede_descargar_informe_boletas_pagos'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Crear el archivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de base de Boletas"

        # Encabezados
        encabezados = [
            "#", "REFERENCIA", "FECHA", "MONTO", "CODIGO", "NOMBRE", "MORA",
            "INTERES", "CAPITAL", "NIT/CUI", "SUMA", "DIFERENCIA", "FACTURA", "RECIBO", "REGISTRO FICTICIO"
        ]
        for col_idx, header in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # Obtener datos
        boletas = self.get_queryset()

        contador = 0

        for idx, boleta in enumerate(boletas, start=2):
            contador += 1
            codigo = '---'
            mora = interes = capital = suma = diferencia = 0
            nombre = 'ELTELAR'
            cui = '---'
            factura = 'NO'
            recibo_p = 'NO'

            recibo = Recibo.objects.filter(pago__id=boleta.id).first()

            if boleta.credit:
                codigo = boleta.credit.codigo_credito
                nombre = boleta.credit.customer_id
                cui = boleta.credit.customer_id.identification_number

            elif boleta.acreedor:
                codigo = boleta.acreedor.codigo_acreedor
                nombre = boleta.acreedor.nombre_acreedor

            elif boleta.seguro:
                codigo = boleta.seguro.codigo_seguro
                nombre = boleta.seguro.nombre_acreedor

            else:
                codigo = boleta.tipo_pago
                if boleta.cliente:
                    nombre = boleta.cliente
                    cui = boleta.cliente.identification_number
                    codigo = boleta.cliente.customer_code

                if boleta.tipo_pago == 'EGRESO':
                    egreso = Egress.objects.filter(numero_referencia=boleta.numero_referencia).first()
                    if egreso:
                        codigo = egreso.codigo_egreso
                        cui = egreso.nit if egreso.nit else '---'
                        nombre = egreso.nombre if egreso.nombre else egreso.observaciones

                elif boleta.tipo_pago == 'INGRESO':
                    ingreso = Income.objects.filter(numero_referencia=boleta.numero_referencia).first()
                    if ingreso:
                        codigo = ingreso.codigo_ingreso

            if recibo:
                mora = recibo.mora_pagada
                interes = recibo.interes_pagado
                capital = recibo.aporte_capital
                suma = mora + interes + capital
                diferencia = boleta.monto - suma
                recibo_p = 'SI'
                factura = 'SI' if recibo.factura else 'NO'

            fila = [
                contador,
                boleta.numero_referencia,
                str(boleta.fecha_emision.date()),
                f'Q {boleta.Fmonto()}',
                codigo,
                str(nombre),
                mora,
                interes,
                capital,
                cui,
                suma,
                diferencia,
                factura,
                recibo_p,
                boleta.registro_ficticio
            ]

            for col_idx, value in enumerate(fila, start=1):
                sheet.cell(row=idx, column=col_idx, value=value)

        # Crear respuesta de descarga
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_boletas_{timestamp}.xlsx"'
        workbook.save(response)
        return response


def report_base_boletas(request):
    # Crear el archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de base de Boletas"
    sucursal = request.session['sucursal_id']

    # Agregar encabezados
    sheet['A1'] = "#"
    sheet['B1'] = "REFERENCIA"
    sheet['C1'] = "FECHA"
    sheet['D1'] = "MONTO"
    sheet['E1'] = "CODIGO"
    sheet['F1'] = "NOMBRE"
    sheet['G1'] = "MORA"
    sheet['H1'] = "INTERES"
    sheet['I1'] = "CAPITAL"
    sheet['J1'] = "NIT/CUI"
    sheet['L1'] = "SUMA"
    sheet['M1'] = "DIFERENCIA"
    sheet['N1'] = "FACTURA"
    sheet['O1'] = "RECIBO"
    

    # Obtener datos de la base de datos, ordenando por status (True primero)
    boletas = Payment.objects.filter(registro_ficticio = False, estado_transaccion='COMPLETADO', sucursal=sucursal).order_by('-id')  # True primero, luego False
    contador = 0

    # Agregar datos al archivo
    for idx, boleta in enumerate(boletas, start=2):  # Comenzar en la fila 2
        contador += 1
        
        codigo = '---'

        mora = 0
        interes = 0
        capital = 0

        cui = '---'

        suma = 0
        diferencia = 0
        nombre = 'ELTELAR'
        cui = '---'

        factura = 'NO'
        recibo_p = 'NO'

        recibo = Recibo.objects.filter(pago__id=boleta.id).first()

        if boleta.credit is not None:
            codigo = boleta.credit.codigo_credito
            nombre = boleta.credit.customer_id

            cui = boleta.credit.customer_id.identification_number

        elif boleta.acreedor is not None:
            codigo = boleta.acreedor.codigo_acreedor
            nombre = boleta.acreedor.nombre_acreedor
            

        elif boleta.seguro is not None:
            codigo = boleta.seguro.codigo_seguro
            nombre = boleta.seguro.nombre_acreedor

        else:
            codigo = boleta.tipo_pago

            if boleta.cliente is not None:
                nombre = boleta.cliente
                cui = boleta.cliente.identification_number
                codigo = boleta.cliente.customer_code
            
            if boleta.tipo_pago == 'EGRESO':
                egreso = Egress.objects.filter(numero_referencia=boleta.numero_referencia).first()
                codigo = egreso.codigo_egreso
                cui = egreso.nit if egreso.nit else '---'
                nombre = egreso.nombre if egreso.nombre else egreso.observaciones

            elif boleta.tipo_pago == 'INGRESO':
                ingreso = Income.objects.filter(numero_referencia=boleta.numero_referencia).first()
                codigo = ingreso.codigo_ingreso


        
        if recibo is not None:
            mora = recibo.mora_pagada
            interes = recibo.interes_pagado
            capital = recibo.aporte_capital

            suma = mora + interes + capital
            diferencia = boleta.monto - suma

            recibo_p = 'SI'

            factura = 'SI' if recibo.factura else 'NO'

        

        sheet[f'A{idx}'] = contador
        sheet[f'B{idx}'] = boleta.numero_referencia
        sheet[f'C{idx}'] = str(boleta.fecha_emision.date())
        sheet[f'D{idx}'] = f'Q {boleta.Fmonto()}'

        sheet[f'E{idx}'] = codigo


        sheet[f'F{idx}'] = str(nombre)
        sheet[f'G{idx}'] = mora
        sheet[f'H{idx}'] = interes
        sheet[f'I{idx}'] = capital
        sheet[f'J{idx}'] = cui
        sheet[f'L{idx}'] = suma
        sheet[f'M{idx}'] = diferencia

        sheet[f'N{idx}'] = factura
        sheet[f'O{idx}'] = recibo_p

       




    # Crear respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="reporte_bolestas_{datetime.now()}.xlsx"'

    # Guardar el archivo en la respuesta
    workbook.save(response)
    return response


