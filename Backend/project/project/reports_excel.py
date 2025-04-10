from apps.financings.models import Banco, Payment, Recibo
from openpyxl import Workbook
from django.http import HttpResponse

def report_filtro_banco(filters):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de Bancos"

    # Agregar encabezados
    sheet['A1'] = "STATUS"
    sheet['B1'] = "FECHA"
    sheet['C1'] = "DESCRIPCION"
    sheet['D1'] = "REFERENCIA"
    sheet['E1'] = "SECUENCIAL"
    sheet['F1'] = "CHEQUE PROPIO/LOCAL/EFECTIVO"
    sheet['G1'] = "DEBITO(-)"
    sheet['H1'] = "CREDITO(+)"
    sheet['I1'] = "SALDO CONTABLE"
    sheet['J1'] = "SALDO DISPONIBLE"
    sheet['L1'] = "MONTO REF COMPARATIVA"
    sheet['M1'] = "TIPO DE PAGO"
    sheet['N1'] = "DESCRIPCION"
    sheet['O1'] = "FECHA"
    sheet['P1'] = "PARA"

    # Obtener datos de la base de datos, ordenando por status (True primero)
    bancos = Banco.objects.filter(filters).order_by('-status')  # True primero, luego False

    # Agregar datos al archivo
    for idx, banco in enumerate(bancos, start=2):  # Comenzar en la fila 2
        numero_referencia = banco.referencia

        # Buscar el objeto Payment asociado
        boleta = Payment.objects.filter(numero_referencia=numero_referencia).first()

        sheet[f'A{idx}'] = banco.status
        sheet[f'B{idx}'] = banco.fecha
        sheet[f'C{idx}'] = banco.descripcion
        sheet[f'D{idx}'] = banco.referencia
        sheet[f'E{idx}'] = banco.secuencial
        sheet[f'F{idx}'] = banco.cheque
        sheet[f'G{idx}'] = banco.debito
        sheet[f'H{idx}'] = banco.credito
        sheet[f'I{idx}'] = banco.saldo_contable
        sheet[f'J{idx}'] = banco.saldo_disponible

        # Agregar datos de boleta si existe
        if boleta:
            sheet[f'L{idx}'] = boleta.monto
            sheet[f'M{idx}'] = boleta.tipo_pago
            sheet[f'N{idx}'] = boleta.descripcion
            sheet[f'O{idx}'] = boleta.fecha_emision.date()
            if boleta.acreedor:
                sheet[f'P{idx}'] = str(boleta.acreedor)
            if boleta.seguro:
                sheet[f'P{idx}'] = str(boleta.seguro)
            if boleta.cliente:
                sheet[f'P{idx}'] = str(boleta.cliente)
            if boleta.credit:
                sheet[f'P{idx}'] = str(boleta.credit)
                
            if boleta.disbursement:
                sheet[f'P{idx}'] = str(boleta.disbursement)




    # Crear respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="reporte_bancos.xlsx"'

    # Guardar el archivo en la respuesta
    workbook.save(response)
    return response

def report_banco(request):
    # Crear el archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de Bancos"

    # Agregar encabezados
    sheet['A1'] = "STATUS"
    sheet['B1'] = "FECHA"
    sheet['C1'] = "DESCRIPCION"
    sheet['D1'] = "REFERENCIA"
    sheet['E1'] = "SECUENCIAL"
    sheet['F1'] = "CHEQUE PROPIO/LOCAL/EFECTIVO"
    sheet['G1'] = "DEBITO(-)"
    sheet['H1'] = "CREDITO(+)"
    sheet['I1'] = "SALDO CONTABLE"
    sheet['J1'] = "SALDO DISPONIBLE"
    sheet['L1'] = "MONTO REF COMPARATIVA"
    sheet['M1'] = "TIPO DE PAGO"
    sheet['N1'] = "DESCRIPCION"
    sheet['O1'] = "FECHA"
    sheet['P1'] = "PARA"

    # Obtener datos de la base de datos, ordenando por status (True primero)
    bancos = Banco.objects.all().order_by('-status')  # True primero, luego False

    # Agregar datos al archivo
    for idx, banco in enumerate(bancos, start=2):  # Comenzar en la fila 2
        numero_referencia = banco.referencia

        # Buscar el objeto Payment asociado
        boleta = Payment.objects.filter(numero_referencia=numero_referencia).first()

        sheet[f'A{idx}'] = banco.status
        sheet[f'B{idx}'] = banco.fecha
        sheet[f'C{idx}'] = banco.descripcion
        sheet[f'D{idx}'] = banco.referencia
        sheet[f'E{idx}'] = banco.secuencial
        sheet[f'F{idx}'] = banco.cheque
        sheet[f'G{idx}'] = banco.debito
        sheet[f'H{idx}'] = banco.credito
        sheet[f'I{idx}'] = banco.saldo_contable
        sheet[f'J{idx}'] = banco.saldo_disponible

        # Agregar datos de boleta si existe
        if boleta:
            sheet[f'L{idx}'] = boleta.monto
            sheet[f'M{idx}'] = boleta.tipo_pago
            sheet[f'N{idx}'] = boleta.descripcion
            sheet[f'O{idx}'] = boleta.fecha_emision.date()
            if boleta.acreedor:
                sheet[f'P{idx}'] = str(boleta.acreedor)
            if boleta.seguro:
                sheet[f'P{idx}'] = str(boleta.seguro)
            if boleta.cliente:
                sheet[f'P{idx}'] = str(boleta.cliente)
            if boleta.credit:
                sheet[f'P{idx}'] = str(boleta.credit)
                
            if boleta.disbursement:
                sheet[f'P{idx}'] = str(boleta.disbursement)




    # Crear respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="reporte_bancos.xlsx"'

    # Guardar el archivo en la respuesta
    workbook.save(response)
    return response

import json
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Q, Sum

def report_pagos(request, filtro_seleccionado, anio, mes, total):
    # Mapeo de filtros válidos
    filtros_validos = {
        'mora_pagada': 'mora_pagada__gt',
        'interes_pagado': 'interes_pagado__gt',
        'aporte_capital': 'aporte_capital__gt',
    }
    filters = Q()
    filters &= Q(fecha__year=anio)
    filters &= Q(fecha__month=mes)
    filters &= Q(pago__registro_ficticio=False)

   

    # Filtrar los reportes
    reportes = None
    
    # Si el filtro seleccionado es válido, aplicar el filtro adicional
    if filtro_seleccionado in filtros_validos:
        filtro_dinamico = {filtros_validos[filtro_seleccionado]: 0}
        reportes = Recibo.objects.filter(filters).filter(**filtro_dinamico)
        

    # Crear el archivo Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Reporte de {filtro_seleccionado}"

    # Agregar encabezados
    sheet['A1'] = f'REPORTE SOBRE {filtro_seleccionado}'
    sheet['F1'] = str(total)
    sheet.append(["#", "FECHA", "CODIGO DEL CREDITO", "CLIENTE", "NO. REFERENCIA", "MONTO PAGADO", "STATUS DEL CREDITO"])

    # Agregar los datos al archivo Excel
    for idx, reporte in enumerate(reportes, start=1):
        monto = None

        # Seleccionar el monto correcto según el filtro
        if filtro_seleccionado == 'mora_pagada':
            monto = reporte.mora_pagada
        elif filtro_seleccionado == 'interes_pagado':
            monto = reporte.interes_pagado
        elif filtro_seleccionado == 'aporte_capital':
            monto = reporte.aporte_capital

        # Agregar los datos a la fila correspondiente
        mensaje = None
        if reporte.pago.credit.estado_aportacion:
            mensaje = 'VIGENTE'
        elif reporte.pago.credit.estado_aportacion is None:
            mensaje = 'SIN APORTACIONES'
        else:
            mensaje = 'EN ATRASO'

        aportacion = mensaje
        s_fecha = 'VIGENTE' if reporte.pago.credit.estados_fechas else 'EN ATRASO'
        sheet.append([
            idx, 
            str(reporte.fecha),
            str(reporte.pago.credit),
            str(reporte.cliente),
            str(reporte.pago.numero_referencia),
            str(monto),
            f"Status de Aportación: {aportacion}, "
            f"Status por Fecha: {s_fecha}"
        ])

    # Crear respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="reportes_sobre_pagos_{filtro_seleccionado}.xlsx"'

    # Guardar el archivo en la respuesta
    workbook.save(response)
    return response
